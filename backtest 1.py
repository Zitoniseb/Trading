import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import warnings

warnings.filterwarnings("ignore", category=FutureWarning)

def flatten_dict(d):
    # Only flatten Series, otherwise return as is
    return {k: (v.values[0] if isinstance(v, pd.Series) and len(v) > 0 else v) for k, v in d.items()}

def force_scalar(val):
    import datetime
    # Exclude non-numeric types before casting to float
    non_numeric_types = (str, bytes, datetime.date, datetime.datetime, datetime.timedelta, np.datetime64, np.timedelta64)
    if isinstance(val, non_numeric_types):
        return None
    # Fast path for numeric scalars
    if is_numeric_scalar(val):
        if isinstance(val, complex):
            return None
        if isinstance(val, (bool, str, bytes, datetime.date, datetime.datetime, datetime.timedelta, np.datetime64, np.timedelta64, pd.Series)):
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None
    # Fast path for 0/1-length Series
    if isinstance(val, pd.Series):
        if len(val) == 0:
            return None
        if len(val) == 1:
            scalar_val = val.iloc[0]
            if isinstance(scalar_val, non_numeric_types) or isinstance(scalar_val, complex):
                return None
            try:
                return float(scalar_val)
            except (TypeError, ValueError):
                return None
        if val.nunique() == 1:
            scalar_val = val.iloc[0]
            if isinstance(scalar_val, non_numeric_types) or isinstance(scalar_val, complex):
                return None
            try:
                return float(scalar_val)
            except (TypeError, ValueError):
                return None
        val = val.values
    # At this point, val could be a numpy array, ExtensionArray, or similar
    if hasattr(val, "__len__") and len(val) == 1:
        scalar_val = val[0]
        if isinstance(scalar_val, non_numeric_types) or isinstance(scalar_val, complex):
            return None
        try:
            return float(scalar_val)
        except (TypeError, ValueError):
            return None
    return None

def is_numeric_scalar(val):
    # Fast path for int/float/numpy numeric
    return isinstance(val, (int, float, np.integer, np.floating))


def calculate_rsi(close_prices, window=14, use_ema=True):
    # Vectorized RSI calculation
    delta = close_prices.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    if use_ema:
        avg_gain = gain.ewm(span=window, adjust=False).mean()
        avg_loss = loss.ewm(span=window, adjust=False).mean()
    else:
        avg_gain = gain.rolling(window).mean()
        avg_loss = loss.rolling(window).mean()
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def process_ticker_strategy(ticker, data):
    try:
        df = data[ticker] if isinstance(data.columns, pd.MultiIndex) else data
        if df is None or df.empty or len(df) < 50:
            return None
        df = df.copy()
        df['Typical_Price'] = (df['High'] + df['Low'] + df['Close']) / 3
        df['TPxVolume'] = df['Typical_Price'] * df['Volume']
        df['VWAP'] = df['TPxVolume'].cumsum() / df['Volume'].cumsum()
        df['EMA_9'] = df['Close'].ewm(span=9, adjust=False).mean()
        df['EMA_20'] = df['Close'].ewm(span=20, adjust=False).mean()
        df['SMA_50'] = df['Close'].rolling(window=50).mean()
        df['RSI'] = calculate_rsi(df['Close'])

        df_10 = df.tail(10)
        current_price = df['Close'].iloc[-1]
        support = df['Low'].tail(20).min()
        resistance = df['High'].tail(20).max()
        rsi = df_10['RSI'].iloc[-1]
        gain_5d = ((current_price - df_10['Close'].iloc[0]) / df_10['Close'].iloc[0]) * 100

        near_support = current_price <= support * 1.05
        risk = current_price - support
        reward = resistance - current_price
        rr_ratio = reward / risk if risk > 0 else None

        decision = "⏳ Watch"
        if rsi < 70 and near_support:
            decision = "✅ BUY"
        elif rsi > 70 or gain_5d > 10:
            decision = "❌ Avoid"

        return {
            "Ticker": ticker,
            "Price": round(current_price, 2),
            "Support": round(support, 2),
            "Resistance": round(resistance, 2),
            "RSI": round(rsi, 2),
            "Gain 5d %": round(gain_5d, 2),
            "Risk": round(risk, 2),
            "Reward": round(reward, 2),
            "RR Ratio": round(rr_ratio, 2) if rr_ratio else None,
            "Near Support": near_support,
            "Decision": decision
        }
    except Exception:
        return None

def backtest_ticker(ticker, signal_date, horizon):
    entry_date = datetime.strptime(signal_date, "%Y-%m-%d")
    end_date = entry_date + timedelta(days=horizon + 5)
    df = yf.download(ticker, start=entry_date - timedelta(days=1), end=end_date, progress=False)
    if df is None or df.empty or entry_date not in df.index:
        return None

    df = df[df.index >= entry_date]
    entry_price = force_scalar(df.iloc[0]['Close'])
    hold_and_wait_price = force_scalar(df.iloc[horizon]['Close'] if len(df) > horizon else df.iloc[-1]['Close'])
    if not is_numeric_scalar(entry_price) or entry_price is None:
        print(f"[ERROR] entry_price is not a numeric scalar: {entry_price}")
        return None
    entry_price = float(entry_price)
    tp_price = entry_price + 4.0
    sl_price = None
    hit_momentum = False
    max_profit = 0.0
    tp_adjust = 6.0
    sl_adjust = 2.5

    exit_price = None
    reason = None
    exit_day = None
    hold_and_wait_price = force_scalar(hold_and_wait_price)
    if is_numeric_scalar(hold_and_wait_price) and hold_and_wait_price is not None:
        hold_and_wait_profit = round(float(hold_and_wait_price) - entry_price, 2)
    else:
        hold_and_wait_profit = None

    def get_scalar_from_df(df, i, col):
        try:
            if col in df.columns:
                val = df.iloc[i][col]
            else:
                val = None
            return force_scalar(val)
        except Exception:
            return None

    for i in range(1, min(horizon + 1, len(df))):
        high = get_scalar_from_df(df, i, 'High')
        close = get_scalar_from_df(df, i, 'Close')
        if not is_numeric_scalar(high) or not is_numeric_scalar(close) or high is None or close is None:
            continue
        date = df.index[i].strftime('%Y-%m-%d')
        profit = high - entry_price if is_numeric_scalar(high) and high is not None and entry_price is not None else None
        # Ensure profit is a real numeric scalar before casting
        if not (is_numeric_scalar(profit) and not isinstance(profit, complex) and profit is not None):
            continue
        profit = float(profit)

        if not hit_momentum and profit >= 3.5:
            tp_adjust = 6.0
            sl_adjust = 2.5
            hit_momentum = True
            max_profit = profit
        elif hit_momentum:
            diff = profit - max_profit
            if diff >= 1.0:
                tp_adjust += 2
                sl_adjust += 1
                max_profit = profit
        if hit_momentum:
            tp_price = entry_price + tp_adjust
            sl_price = entry_price + sl_adjust

        if not is_numeric_scalar(tp_price) or tp_price is None:
            continue
        if not (sl_price is None or (is_numeric_scalar(sl_price) and sl_price is not None)):
            continue

        if is_numeric_scalar(high) and is_numeric_scalar(tp_price) and high is not None and tp_price is not None and high >= tp_price:
            exit_price = tp_price
            reason = "TP Hit"
            exit_day = date
            break

        if sl_price is not None and is_numeric_scalar(close) and is_numeric_scalar(sl_price) and close is not None and sl_price is not None and close <= sl_price:
            exit_price = close
            reason = "SL Hit"
            exit_day = date
            break

    exit_price = force_scalar(exit_price)
    if exit_price is None:
        exit_price = df.iloc[min(horizon, len(df) - 1)]['Close']
        reason = "Timeout"
        exit_day = df.index[min(horizon, len(df) - 1)].strftime('%Y-%m-%d')

    entry_price = force_scalar(entry_price)
    exit_price = force_scalar(exit_price)
    hold_and_wait_price = force_scalar(hold_and_wait_price)
    if is_numeric_scalar(exit_price) and is_numeric_scalar(entry_price) and exit_price is not None and entry_price is not None:
        actual_profit = round(float(exit_price) - float(entry_price), 2)
    else:
        actual_profit = None

    return {
        "Ticker": ticker,
        "Signal Date": signal_date,
        "Entry Price": round(entry_price, 2) if is_numeric_scalar(entry_price) and entry_price is not None else None,
        "Exit Price (TP/SL)": round(exit_price, 2) if is_numeric_scalar(exit_price) and exit_price is not None else None,
        "Final Profit (€)": actual_profit,
        "Exit Reason": reason,
        "Exit Date": exit_day,
        "Hold & Wait Price": round(hold_and_wait_price, 2) if is_numeric_scalar(hold_and_wait_price) and hold_and_wait_price is not None else None,
        "Hold & Wait Profit (€)": hold_and_wait_profit
    }

# === Main Flow ===
tickers = [
    "005930.KS", "0700.HK", "0883.HK", "0939.HK", "2330.TW", "600519.SS", "6861.T", "7203.T", "AAPL",
    "ABBV", "ABT", "ACN", "ADBE", "AEIS", "AMGN", "AMZN", "ASML", "AVGO", "AZN", "BA", "BAC", "BABA",
    "BHP.AX", "BP", "BRK-B", "CBA.AX", "CAT", "COIN", "COST", "CRM", "CSCO", "CVX", "DAL", "DIS",
    "DTE.DE", "ENEL.MI", "GOOG", "GOOGL", "HD", "HON", "HSBA.L", "IBM", "INTC", "JNJ", "JPM", "KO", "LLY",
    "LOW", "MA", "MBG.DE", "MC.PA", "MCD", "MELI", "META", "MMM", "MRK", "MSFT", "NESN.SW", "NFLX", "NKE",
    "NVDA", "ORCL", "PEP", "PFE", "PG", "PYPL", "QCOM", "RIO.AX", "RIO.L", "ROG.SW", "RTX", "SAN.PA",
    "SAP.DE", "SBUX", "SHEL", "TCEHY", "TMO", "TSLA", "UBER", "ULVR.L", "UNH", "UPS", "V", "WFC", "WMT",
    "XOM"
]

today = datetime.today()
signal_dates = [(today - timedelta(days=5 + i * 7)).strftime('%Y-%m-%d') for i in range(9)]

wb = Workbook() 
if wb.active is not None:
    wb.remove(wb.active)
pdf_rows = []

for signal_date in signal_dates:
    signal_day = datetime.strptime(signal_date, "%Y-%m-%d")
    data = yf.download(tickers, start=signal_day - timedelta(days=90), end=signal_day, group_by="ticker", progress=False)
    with ThreadPoolExecutor(max_workers=12) as executor:
        strategy_results = list(executor.map(lambda t: process_ticker_strategy(t, data), tickers))
    strategy_df = pd.DataFrame([r for r in strategy_results if r is not None])
    buys = strategy_df[strategy_df["Decision"].str.contains("BUY", case=False)]

    for horizon in ([5] if signal_date == signal_dates[0] else [5, 10]):
        rows = []
        for ticker in buys["Ticker"]:
            result = backtest_ticker(ticker, signal_date, horizon)
            if result:
                rows.append(result)
                pdf_rows.append(result)

        if rows:
            df_result = pd.DataFrame(rows)
            if not df_result.empty:
                if any(df_result.applymap(lambda x: isinstance(x, pd.Series)).any()):
                    df_result = df_result.apply(flatten_dict, axis=1).apply(pd.Series)
                ws = wb.create_sheet(title=f"{signal_date}_{horizon}d")
                for r in dataframe_to_rows(df_result, index=False, header=True):
                    ws.append(r)

if not wb.sheetnames:
    ws = wb.create_sheet(title="No Results")
    ws.append(["No backtest results for the given dates."])

wb.save("buy_signal_backtests.xlsx")

df_pdf = pd.DataFrame(pdf_rows)
if not df_pdf.empty:
    if any(df_pdf.applymap(lambda x: isinstance(x, pd.Series)).any()):
        df_pdf = df_pdf.apply(flatten_dict, axis=1).apply(pd.Series)

if not df_pdf.empty:
    summary = {
        "Total Trades": len(df_pdf),
        "Win Rate": (df_pdf["Final Profit (€)"] > 0).mean() * 100,
        "Avg Profit (€)": df_pdf["Final Profit (€)"].mean(),
        "Median Profit (€)": df_pdf["Final Profit (€)"].median(),
        "Max Profit (€)": df_pdf["Final Profit (€)"].max(),
        "Min Profit (€)": df_pdf["Final Profit (€)"].min(),
        "Avg Hold&Wait (€)": df_pdf["Hold & Wait Profit (€)"].mean(),
    }

    summary["Best Trade"] = df_pdf.loc[df_pdf["Final Profit (€)"].idxmax()].to_dict()
    summary["Worst Trade"] = df_pdf.loc[df_pdf["Final Profit (€)"].idxmin()].to_dict()

    print("\n=== Backtest Summary ===")
    for k, v in summary.items():
        if isinstance(v, dict):
            print(f"{k}: {v['Ticker']} | Profit: {v['Final Profit (€)']} | Exit: {v['Exit Reason']} on {v['Exit Date']}")
        else:
            print(f"{k}: {v:.2f}" if isinstance(v, float) else f"{k}: {v}")

    with pd.ExcelWriter("buy_signal_backtests.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Summary")

    melted = df_pdf.melt(id_vars=["Ticker"], value_vars=["Final Profit (€)", "Hold & Wait Profit (€)"])
    plt.figure(figsize=(12, 6))
    # Use seaborn's efficient barplot with precomputed data, avoid unnecessary recalculation
    sns.barplot(data=melted, x="Ticker", y="value", hue="variable", dodge=True, ci=None)
    plt.title("Strategy vs Hold-and-Wait Profit (€)")
    plt.ylabel("Profit (€)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig("buy_signal_backtests_chart.pdf")
    plt.close()  # Free memory after saving the figure

print("\n✅ All files generated: Excel + PDF")
